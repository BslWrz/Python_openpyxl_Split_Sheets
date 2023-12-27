#   @Author BslWrz
#   @Date 2023/12/26

import openpyxl
import copy


def split_worksheets(input_file, output_path):
    wb = openpyxl.load_workbook(input_file, data_only=True)
    sheet_names = wb.sheetnames

    # 存储合并单元格范围和列宽
    merged_cells_dic = {}
    column_dimensions_dic = {}
    for sheet_name in sheet_names:
        source_sheet = wb[sheet_name]
        merged_cells_dic[sheet_name] = source_sheet.merged_cells
        column_dimensions_dic[sheet_name] = source_sheet.column_dimensions

    wb.close()

    # 只读模式节约内存
    wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
    for sheet_name in sheet_names:
        # 创建一个新的工作簿
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # 移除默认创建的工作表

        # 复制当前工作表的数据
        source_sheet = wb[sheet_name]
        new_sheet = new_wb.create_sheet(title=sheet_name)

        # 设置列宽
        for column in column_dimensions_dic[sheet_name]:
            column_width = column_dimensions_dic[sheet_name][column].width
            new_sheet.column_dimensions[column].width = column_width

        # for row in source_sheet.iter_rows(values_only=True):
        #     new_sheet.append(row)
        for row in source_sheet.iter_rows():
            for cell in row:
                # 获取目标单元格，将源单元格的值复制到目标单元格
                # print(type(cell))
                if isinstance(cell, openpyxl.cell.read_only.EmptyCell):
                    continue
                # print(get_column_letter(cell.column), cell.column, cell.value)
                target_cell = new_sheet[cell.coordinate]
                target_cell.value = cell.value

                # 复制源单元格的样式到目标单元格
                # target_cell._style = copy.copy(cell._style)
                target_cell.font = copy.copy(cell.font)  # 字体
                target_cell.border = copy.copy(cell.border)  # 边框
                target_cell.fill = copy.copy(cell.fill)  # 填充颜色
                target_cell.number_format = copy.copy(cell.number_format)  # 数字格式
                # target_cell.protection = copy.copy(cell.protection)
                target_cell.alignment = copy.copy(cell.alignment)  # 对齐

        # 处理源单元格的合并单元格
        for merged_cell in merged_cells_dic[sheet_name]:
            range_string = str(merged_cell)
            new_sheet.merge_cells(range_string)

        # 保存新的工作簿
        new_file_name = f"{output_path}/{sheet_name}.xlsx"
        new_wb.save(new_file_name)

    wb.close()


# 测试代码
split_worksheets('input.xlsx', 'output')
